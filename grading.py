from os import system, name
from openpyxl import Workbook
import numpy as np

grade = {'A+': 90, 'A' : 80, 'A-' : 70, 'B+' : 65, 'B' : 60, 'C+' : 55, 'C' : 50, 'D' : 45, 'E' : 40, 'G' : 0}

core_subject1 = ["Bahasa Melayu", "English", "Mathematics", "History"]
core_subject2 = ["Bahasa Melayu", "English", "Mathematics", "History", "Pendidikan Islam/Pendidikan Moral"]
core_subject3 = ["Bahasa Melayu", "English", "Mathematics", "History", "Science", "Pendidikan Islam/Pendidikan Moral"]

preset_subject1 = ["Physics","Chemistry","Biology","Additional Mathematics"]

elective_subject = ["Bahasa Arab", "Pendidikan Syariah Islamiah", "Pendidikan Al-Quran dan As-Sunnah" ,"Bahasa Cina", "Bahasa Tamil", "Physics", "Chemistry", "Biology", "Additional Mathematics", "Account", "Computer Science", "Economy", "Art", "Business", "Geography"]
additional_subject = []

cli_subjects = []
cli_subjects_mark = []
cli_subjects_grade = []
all_subjects_mark = 0
cli_overall_marks = 0
cli_overall_percentage = 0.0
cli_result = ""


def clear():
    # for windows
    if name == 'nt':
        _ = system('cls')
 
    # for mac and linux(here, os.name is 'posix')
    else:
        _ = system('clear')


def query(qu):
    while True:
        cli_input = input(qu)
        Fl = cli_input[0].lower()
        if cli_input == '' or not Fl in ['y','n']:
            print('Please respond with [Y]yes or [n]no!')
        else:
            return Fl
            break


def _select_core_subject():
    clear()
    while True:
        cli_input = input("\n"
                "  +  Core Subject 1 "+str(core_subject1)+", type [0]\n"
                "  +  Core Subject 2 "+str(core_subject2)+", type [1]\n"
                "  +  Core Subject 3 "+str(core_subject3)+", type [2]\n\n"
                "> Select your core subjects: "
                )
        if cli_input == '' or not [0,1,2]: 
            print("Please respond correctly!")
        else:
            break
    c_subs = [core_subject1, core_subject2, core_subject3]
    for i in range(0,len(c_subs[int(cli_input)])):
        cli_subjects.append(c_subs[int(cli_input)][i])
    clear()
    _select_elective_subject()


def _select_elective_subject():
    while True:
        cli_input = input("\n"
            "  +  Elective Subject 1 "+str(preset_subject1)+", type [0]\n"
            "  +  Choose your own elective subject, type [1]\n\n"
            "> Select your elective subject: "
            )
        if cli_input == '' or not [0,1]:
            print("Please respond correctly!")
        else:
            break
    ps_subs = [preset_subject1]
    if int(cli_input) == 1:
        clear()
        _select_specific_elective_subject()
    else:
        for i in range(0,len(ps_subs[int(cli_input)])):
            cli_subjects.append(ps_subs[int(cli_input)][i])
    clear()
    _select_subject_confirmation()


def _select_specific_elective_subject():
    for i in range(0, len(elective_subject)):
        qinput = "\n> Add "+str(elective_subject[i])+" as your subject? [Y/n] "
        reply = query(qinput)
        if reply == "y":
            cli_subjects.append(elective_subject[i])
        elif reply=="n":
            pass


def _select_subject_confirmation():
    qinput = "\n> Do you confirm this is your subject? "+ str(cli_subjects)+ ", [Y/n] "
    reply = query(qinput)
    if reply == "y":
        pass
    elif reply =="n":
        exit()
    _subject_marking()


def _subject_marking():
    clear()
    for i in range(0, len(cli_subjects)):
        while True:
            cli_input = input("\nWhat is the examination mark for ["+cli_subjects[i]+"]?\n> ")
            if not cli_input.isnumeric() or int(cli_input)>100:
                print("Please respond correctly, and maximum mark is only 100!")
            else:
                cli_subjects_mark.append(int(cli_input))
                break
    _result()


def _subject_grading():
    for i in cli_subjects_mark:
        for j in grade:
            if i >= grade[j]:
                cli_subjects_grade.append(j)
                break


def _overall_mark():
    global all_subjects_mark
    global cli_overall_marks
    x = 0
    all_subjects_mark = len(cli_subjects) * 100
    for i in cli_subjects_mark:
        x += i
    cli_overall_marks = x
    print("Overall Marks:", str(cli_overall_marks), "/", all_subjects_mark)


def _overall_percentage():
    global cli_overall_percentage
    cli_overall_percentage = (cli_overall_marks/all_subjects_mark) * 100
    print("Percentage:", str(cli_overall_percentage) + "%")


def _number_of_subject():
    print("Number of Subject: "+str(len(cli_subjects))+"\n")


def _result():
    global cli_result
    clear()
    print("\nEXAMINATION RESULT:\n")
    _subject_grading()
    _number_of_subject()
    for i in range(0,len(cli_subjects)):
        print(cli_subjects[i]+" "+str(cli_subjects_mark[i])+" "+cli_subjects_grade[i])
    print("\nOverall Grades: "+str(cli_subjects_grade))
    _overall_mark()
    _overall_percentage()
    x = "Result: "
    if cli_overall_percentage > grade["E"] and cli_subjects_mark[3] > grade["E"] and cli_subjects_mark[0] > grade["E"]: # check if History and BM passed or not
        cli_result = "PASSED"
    else:
        cli_result = "FAILED"
    print(x, cli_result)
    reply = query("\n> Would you like to create a spreadsheet? [Y/n] ")
    if reply == "y":
        _create_result_spreadsheet()
    elif reply=="n":
        exit()


def _create_result_spreadsheet():
    wb = Workbook()
    ws = wb.active

    wbname = input("\n> Enter a filename for your spreadsheet: ")

    clmn1 = np.array(["Overall Grades", "Overall Marks", "Percentage", "Result"])
    clmn2 = np.array([str(cli_subjects_grade), str(str(cli_overall_marks) + "/" + str(all_subjects_mark)), cli_overall_percentage, cli_result])
    clmn3 = cli_subjects_grade

    clmn1 = np.insert(clmn1, 0 , cli_subjects)
    clmn2 = np.insert(clmn2, 0 , cli_subjects_mark)
    
    for i in range(len(clmn1)):
        ws.cell(row=i+1, column=1, value=clmn1[i])
    for i in range(len(clmn2)):
        ws.cell(row=i+1, column=2, value=clmn2[i])
    for i in range(len(clmn3)):
        ws.cell(row=i+1, column=3, value=clmn3[i])

    wb.save(wbname+".xlsx")


def main():
    _select_core_subject()


main()
